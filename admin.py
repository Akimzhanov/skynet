from django.contrib import admin 
from datetime import datetime
from rangefilter.filters import DateRangeFilter, DateTimeRangeFilter, NumericRangeFilter
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Sum,FloatField, F
from django.db.models.functions import Cast
from django.db.models import Q
from django.contrib import messages
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from .models import Bitrix_1C,Montajnik_1C,Change_raz
from django.db.models.functions import Coalesce
from django.utils.translation import gettext as T_E
from fast_bitrix24 import BitrixAsync
import asyncio, time
import locale
from num2words import num2words


class CustomFilter(admin.SimpleListFilter):
    title = 'Разряды монтажников'  # Название фильтра, отображаемое в интерфейсе
    parameter_name = 'custom_filter'  # Параметр фильтра, который будет использоваться в URL

    def lookups(self, request, model_admin):
        values = Change_raz.objects.values_list('chislo', flat=True)
        return [(value, value) for value in values]

    def queryset(self, request, queryset):
        return None


async def mont_raz_1_2(deal_id):
    webhook = "" # токен с битрикс24
    b = BitrixAsync(webhook)
    method = 'crm.deal.get'
    params = {
        'ID': deal_id
    }

    test = await b.call(method, params, raw=True)
    tv = test['result']['UF_CRM_1674021434343']
    wtv = test['result']['UF_CRM_1676284476112']
    ovk = test['result']['UF_CRM_1674024846']
    money = 1000
    tv_list = ['Приставка Imaqliq G-box ВП', 'Приставка Imaqliq G-box выкуп', 'Приставка Imaqliq Q-box ВП', 'Приставка Imaqliq Q-box выкуп']

    if tv in tv_list:
        money += 300
    elif tv == 'Приложение':
        money += 500

    if wtv in tv_list:
        money += 300
    elif wtv == 'Приложение':
        money += 800

    if float(ovk) > 350:
        money += 250

    return money

async def mont_raz_3(deal_id):
    webhook = "" # токен с битрикс24
    b = BitrixAsync(webhook)
    method = 'crm.deal.get'
    params = {
        'ID': deal_id
    }


    test = await b.call(method, params, raw=True)
    tv = test['result']['UF_CRM_1674021434343']
    wtv = test['result']['UF_CRM_1676284476112']
    ovk = test['result']['UF_CRM_1674024846']
    money = 750
    tv_list = ['Приставка Imaqliq G-box ВП', 'Приставка Imaqliq G-box выкуп', 'Приставка Imaqliq Q-box ВП', 'Приставка Imaqliq Q-box выкуп']

    if tv in tv_list:
        money += 300
    elif tv == 'Приложение':
        money += 500

    if wtv in tv_list:
        money += 300
    elif wtv == 'Приложение':
        money += 800

    if int(ovk) > 350:
        money += 250

    return money


async def tehpod(deal_id):
    webhook = "" # токен с битрикс24
    b = BitrixAsync(webhook)
    method = 'crm.deal.get'
    params = {
        'ID': deal_id
    }


    test = await b.call(method, params, raw=True)
    tv = test['result']['UF_CRM_1674021434343']
    pri_comment = test['result']['UF_CRM_1678770251758']

    money = 450
    tv_list = ['Приставка Imaqliq G-box ВП', 'Приставка Imaqliq G-box выкуп', 'Приставка Imaqliq Q-box ВП', 'Приставка Imaqliq Q-box выкуп']

    if tv in tv_list and 'Замена ТВ' not in pri_comment:
        money += 300
    elif tv == 'Приложение' and 'Замена ТВ' not in pri_comment:
        money += 500


    return money



async def tv_ust(deal_id):
    webhook = "" # токен с битрикс24
    b = BitrixAsync(webhook)
    method = 'crm.deal.get'
    params = {
        'ID': deal_id
    }

    test = await b.call(method, params, raw=True)
    tv = test['result']['UF_CRM_1674021434343']
    money = 0
    tv_list = ['Приставка Imaqliq G-box ВП', 'Приставка Imaqliq G-box выкуп', 'Приставка Imaqliq Q-box ВП', 'Приставка Imaqliq Q-box выкуп']

    if tv in tv_list:
        money += 300
    elif tv == 'Приложение':
        money += 500


    return money



@admin.register(Bitrix_1C)
class ReportAdmin(admin.ModelAdmin):
    list_display = ['bx_id', 'id_mont', 'get_fio_montajnik', 'addres', 'ls_abon', 'money', 'date_tg']
    list_filter = [('date_tg', DateRangeFilter), CustomFilter]
    search_fields = ['id_mont', 'montajnik__fio_mont']
    list_per_page = 200

    def get_fio_montajnik(self, obj):
        if obj.id_mont:
            montajnik = Montajnik_1C.objects.filter(id_mont=obj.id_mont).first()
            if montajnik:
                return montajnik.fio_mont
        return None

    def export_bitrix_data_to_excel(modeladmin, request, queryset):
        razryd = request.GET.get('custom_filter')
        queryset = queryset.filter(status='1')
        
        wb = Workbook()
        ws = wb.active
    
        if not queryset:
            response = HttpResponse("No data to export.")
            return response

        te1 = 'Сумма \n взымаемая \n с абонента(сом)'
        te2 = 'Стоимость \n выполненных \n работ(сом)'
        te3 = 'Наименование \n работы'
        columns = ['Адрес', 'ЛС', te1, te2, te3]
        second_table_columns = ['Наименование оборудования и материалов', 'Остаток', 'Единица измерения']
        akt = 'AKT №1'
        works_inp ='Сдачи-приемки выполненных работ к заказу №1'
        locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
        today = datetime.now().strftime('%d %B %Y года')
        text_1 = '    Исполнитель и Заказчик настоящим Актом произвели сдачу - приемку выполненных Работ на общую сумму: (деньги за работу монажника).\n    Исполнителем выполнены следующие работына объектах заказчика: подключение абонентов к сети абонентского доступа ОсОО <<Скайнет Телеком>> \nс установкой необходимых крепежных материалов и оборудования согласно перечню выполненных работ.'
        text_2 = 'Перечень выполненных работ:'
        visa_customer = 'Виза ответ. Исполнителя Заказчика________________'
        executor = 'Исполнитель__________________'
        list_executor = 'Перечень материалов, потраченных исполнителем:'
        text_3 = '    Работы выполнены надлежащим образомб сданы-приняты комплектно, объязательств сторон согласно Договору выполнены надлежащим образом \nи в полном объёме, качеств произведенных Работ соответствует требованиям Договора. \n    Претензии у Сторон отсутствуют. '



    
        montajnik = Montajnik_1C.objects.filter(id_mont=queryset[0].id_mont).first()

        if montajnik:
            owner = montajnik.fio_mont

        else:
            owner = ''


        row_start = 9
        font = Font(size=11)
        font2 = Font(size=12)
        ws.cell(row=1, column=3, value=akt).font=font
        ws.cell(row=3, column=2, value=works_inp).font=font
        ws.cell(row=5, column=5, value=today).font=font
        ws.cell(row=9, column=1, value=text_2)

        ws.row_dimensions[11].height = 40
        for idx, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_start+2, column=idx, value=column_title)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            if idx == 1:  # Первый столбец
                ws.column_dimensions[get_column_letter(idx)].width = 60
            elif idx == 2:  # Первый столбец
                ws.column_dimensions[get_column_letter(idx)].width = 15
            else:  # Остальные столбцы
                ws.column_dimensions[get_column_letter(idx)].width = 17

        mont_sum = 0 
        podkl = queryset.filter(dismantling='0')
        teh = queryset.filter(dismantling='2')
        tv = queryset.filter(dismantling='3')
        list_work = [podkl, teh, tv]
        length_table_all = 0
        all_sum = 0
        onu = 0
        ovk = 0
        patchcord = 0
        odf = 0
        utp_lenght = 0
        kronshtein = 0
        connecter = 0
        snr = 0
        tp_link_841 = 0
        tp_link_844 = 0
        tp_link_c24 = 0
        tp_link_c54 = 0
        tv_q_box = 0
        tv_g_box = 0
        for datas in list_work:
            for row_num, data in enumerate(datas, 4):
                deal_id = data.bx_id
                if data.dismantling == '0':
                    if int(razryd) == 1 or int(razryd) == 2:
                        tip_work = 'Подключение'
                        mont_money = asyncio.run(mont_raz_1_2(deal_id))
                        mont_sum += mont_money
                        money_value = float(data.money) if data.money else 0
                        row_data = [data.addres, data.ls_abon, money_value, mont_money, tip_work]
                        ws.append(row_data)
                    elif int(razryd) == 3:
                        tip_work = 'Подключение'
                        mont_money = asyncio.run(mont_raz_3(deal_id))
                        mont_sum += mont_money
                        money_value = float(data.money) if data.money else 0
                        row_data = [data.addres, data.ls_abon, money_value, mont_money, tip_work]
                        ws.append(row_data)
                elif data.dismantling == '2':
                    tip_work = 'Техпод'
                    mont_money = asyncio.run(tehpod(deal_id))
                    mont_sum += mont_money

                    money_value = float(data.money) if data.money else 0
                    row_data = [data.addres, data.ls_abon, money_value, mont_money, tip_work]
                    ws.append(row_data)

                elif data.dismantling == '3':
                    tip_work = 'Установка ТВ'
                    mont_money = asyncio.run(tv_ust(deal_id))
                    mont_sum += mont_money

                    money_value = float(data.money) if data.money else 0
                    row_data = [data.addres, data.ls_abon, money_value, mont_money, tip_work]
                    ws.append(row_data)

            ws.append(["", ""])

            total_money = sum(float(data.money) if data.money else 0 for data in datas)
            all_sum += total_money
            length_table = len(datas)
            length_table_all += length_table
            for data in datas:
                onu += float(data.onu) if data.onu else 0
                ovk += float(data.ovk1) if data.ovk1 else 0
                patchcord += 1 if str(data.patchcord) == '000000754' else 0
                odf += float(data.odf) if data.odf else 0
                utp_lenght += float(data.utp_lenght) if data.utp_lenght else 0
                kronshtein += float(data.kronshtein) if data.kronshtein else 0
                connecter += float(data.connecter) if data.connecter else 0
                snr += 1 if str(data.router) == '000000669' else 0
                tp_link_841 += 1 if str(data.router) == '000000014' else 0
                tp_link_844 += 1 if str(data.router) == '000001373' else 0
                tp_link_c24 += 1 if str(data.router) == '000001376' else 0
                tp_link_c54 += 1 if str(data.router) == '000001396' else 0
                tv_q_box += 1 if str(data.tv) == '000000850' else 0
                tv_g_box += 1 if str(data.tv) == '000001345' else 0

        second_table_values = [
            ("Абонентский терминал ONU BDCOM GEPON", onu),
            ("Кабель оптический Alfa Mile Flex FTTx 1 вол", ovk),
            ("Патчкорд оптический", patchcord),
            ("Разетка оптическая SNR-FTB_02S-ODF", odf),
            ("Кабель КСВПВ", utp_lenght),
            ("Кронштейн", kronshtein),
            ("Конектор RJ-45-5e", connecter),
            ("SNR", snr),
            ("TP-Link 841", tp_link_841),
            ("TP-Link 844", tp_link_844),
            ("TP-Link c24", tp_link_c24),
            ("TP-Link c54", tp_link_c54),
            ("Приставка Imaqliq Q-box", tv_q_box),
            ("Приставка Imaqliq G-box", tv_g_box)
        ]
        second_table_start_row = length_table_all + row_start + 14  # Начальная строка для второй таблицы
    
        ws.append(["Общая сумма:", "", all_sum, mont_sum])
        ws.append(["","","",""])
        ws.append([visa_customer, "", "", executor])
        ws.append(["","","",""])

        ws.append([list_executor, ""])
        ws.append(["","","",""])

        ws.append(second_table_columns)
        for idx, (param, value) in enumerate(second_table_values, second_table_start_row):
            ws.cell(row=idx, column=1, value=param)
            ws.cell(row=idx, column=2, value=value)
    
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append([text_3])
        ws.row_dimensions[ws.max_row].height = 60
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append([visa_customer, "", "", executor])

        montajnik = Montajnik_1C.objects.filter(id_mont=queryset[0].id_mont).first()
        testyy = str(montajnik.fio_mont).replace(' ', '_')
        text_number = num2words(mont_sum, lang='ru')
        text_1 = f'    Исполнитель и Заказчик настоящим Актом произвели сдачу - приемку выполненных Работ на общую сумму: {mont_sum} сом({text_number}).\n    Исполнителем выполнены следующие работына объектах заказчика: подключение абонентов к сети абонентского доступа ОсОО <<Скайнет Телеком>> \nс установкой необходимых крепежных материалов и оборудования согласно перечню выполненных работ.'

        ws.cell(row=7, column=1, value=text_1)
        ws.row_dimensions[7].height = 70
        ws.column_dimensions[get_column_letter(7)].width = 40     
        owner_slug = str(time.strftime('%d.%m.%Y %H:%M:%S'))
        response = HttpResponse(content_type='application/ms-excel; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename={owner_slug}.xlsx'
        wb.save(response)
    
        return response
    
    export_bitrix_data_to_excel.short_description = T_E("Экспортировать выбранные записи в Excel")
    actions = [export_bitrix_data_to_excel]








